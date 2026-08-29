from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ats_scan.models.run import RunResult
from ats_scan.report.xlsx import XlsxWriter


@pytest.fixture
def xlsx_path(tmp_path: Path, sample_run: RunResult) -> Path:
    writer = XlsxWriter()
    result = writer.write(sample_run, tmp_path)
    assert result.ok
    return result.value


def test_xlsx_sheets_exist(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    assert set(wb.sheetnames) == {"summary", "dimensions", "diagnostics"}


def test_xlsx_summary_rows(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["summary"]
    assert ws.cell(row=1, column=1).value.startswith("This output is decision support only")
    header = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
    assert "candidate_id" in header
    assert "composite" in header
    data = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
    assert data[header.index("candidate_id")] == "c_abc123"
    assert data[header.index("composite")] == 87.06


def test_xlsx_dimensions_long_format(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["dimensions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "c_abc123" and row[1] == "S1" and row[2] == 88.4 for row in rows)
    assert sum(1 for row in rows if row[0] == "c_abc123") == 10


def test_xlsx_diagnostics_sheet(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["diagnostics"]
    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    data = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
    assert data[header.index("candidate_id")] == "c_abc123"
    assert data[header.index("confidence")] == 0.91


def test_xlsx_conditional_formatting_on_composite(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb["summary"]
    composite_col = next(
        col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=col).value == "composite"
    )
    composite_letter = get_column_letter(composite_col)
    ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    assert any(composite_letter in r for r in ranges)


def test_xlsx_atomic_write(tmp_path: Path, sample_run: RunResult) -> None:
    writer = XlsxWriter()
    writer.write(sample_run, tmp_path)
    assert not list(tmp_path.glob("*.tmp"))
