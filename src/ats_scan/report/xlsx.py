from __future__ import annotations

import os
from pathlib import Path
from typing import ClassVar

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import ColorScaleRule  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from ats_scan.models.common import StageResult
from ats_scan.models.run import RunResult
from ats_scan.models.scoring import ScoreCard
from ats_scan.protocols import ReportWriter
from ats_scan.report._helpers import (
    DECISION_SUPPORT_BANNER,
    _candidate_file,
    _candidate_name,
    matched_required,
    missing_required,
    relevant_years,
    semicolon_join,
    sub_score_value,
)


class XlsxWriter(ReportWriter):
    """Write ``scores.xlsx`` with summary, dimensions and diagnostics sheets.

    TRD §9.1 / FR-903: workbook with three sheets and conditional formatting on
    the composite column.
    """

    artefact: ClassVar[str] = "scores.xlsx"

    def write(self, run: RunResult, out_dir: Path) -> StageResult[Path]:
        path = out_dir / self.artefact
        wb = Workbook()
        wb.remove(wb.active)

        self._write_summary(wb, run)
        self._write_dimensions(wb, run)
        self._write_diagnostics(wb, run)

        tmp = path.with_suffix(path.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, path)
        return StageResult(value=path)

    def _write_summary(self, wb: Workbook, run: RunResult) -> None:
        ws: Worksheet = wb.create_sheet("summary")
        header = [
            "rank",
            "candidate_id",
            "file",
            "name",
            "composite",
            "band",
            "selected",
            "eligible",
            "confidence",
            "S1",
            "S2",
            "S3",
            "S4",
            "S5",
            "S6",
            "S7",
            "S8",
            "S9",
            "S10",
            "matched_required",
            "missing_required",
            "relevant_years",
            "flags",
            "reason_codes",
            "explanation",
        ]

        ws.append([DECISION_SUPPORT_BANNER])
        ws.append(header)
        for cell in ws[2]:
            cell.font = Font(bold=True)

        for card in run.scorecards:
            ws.append(self._summary_row(card, run))

        self._apply_composite_formatting(ws, header.index("composite") + 1, len(run.scorecards) + 2)
        self._autofit_columns(ws, header)

    def _summary_row(self, card: ScoreCard, run: RunResult) -> list[object]:
        return [
            card.rank if card.rank is not None else "",
            card.candidate_id,
            _candidate_file(card, run),
            _candidate_name(card, run),
            card.composite if card.composite is not None else "",
            card.band.value if card.band else "",
            "true" if card.selected else "false",
            "true" if card.eligible else "false",
            card.confidence if card.confidence is not None else "",
            *[
                sub_score_value(card, f"S{i}") if sub_score_value(card, f"S{i}") is not None else ""
                for i in range(1, 11)
            ],
            matched_required(card),
            missing_required(card),
            relevant_years(card),
            semicolon_join(card.flags),
            semicolon_join(card.reason_codes),
            card.explanation,
        ]

    def _apply_composite_formatting(self, ws: Worksheet, col: int, last_row: int) -> None:
        col_letter = get_column_letter(col)
        rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{col_letter}3:{col_letter}{last_row}", rule)

    def _write_dimensions(self, wb: Workbook, run: RunResult) -> None:
        ws: Worksheet = wb.create_sheet("dimensions")
        ws.append(["candidate_id", "dimension", "value", "notes"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for card in run.scorecards:
            for dimension in (f"S{i}" for i in range(1, 11)):
                value = sub_score_value(card, dimension)
                sub = card.sub_scores.get(dimension)
                notes = ";".join(sub.notes) if sub is not None else ""
                ws.append([card.candidate_id, dimension, value if value is not None else "", notes])
        self._autofit_columns(ws, ["candidate_id", "dimension", "value", "notes"])

    def _write_diagnostics(self, wb: Workbook, run: RunResult) -> None:
        ws: Worksheet = wb.create_sheet("diagnostics")
        ws.append(["candidate_id", "confidence", "flags", "reason_codes", "knockouts"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for card in run.scorecards:
            ko_summary = ";".join(f"{ko.id}={ko.verdict}" for ko in card.knockout_results)
            ws.append(
                [
                    card.candidate_id,
                    card.confidence if card.confidence is not None else "",
                    semicolon_join(card.flags),
                    semicolon_join(card.reason_codes),
                    ko_summary,
                ]
            )
        self._autofit_columns(
            ws, ["candidate_id", "confidence", "flags", "reason_codes", "knockouts"]
        )

    def _autofit_columns(self, ws: Worksheet, header: list[str]) -> None:
        for col_idx, title in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(title) + 2)
